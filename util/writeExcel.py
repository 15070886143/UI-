#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2019/2/25 16:52
# @Author  : writeExcel
import openpyxl
from openpyxl.styles import Border,Side,Font
from DataDrivenFrameWork.appModels import readExcel
import time
# class writeexcel():
#     def __init__(self,expath):
#         self.expath = expath
#         self.wb = openpyxl.load_workbook(self.expath)
#         self.ws = self.wb.active#激活sheet
#
#     def write(self,rowns,colns,value):
#         self.ws.cell(rowns,colns).value = value
#         # self.ws.cell(10,10,'woaini')
#         self.wb.save(self.expath)
class write():
    def __init__(self):
        pass
    def writecell(self,sheet,content,coordinate= None,rowno= None,colsno = None,style = None):
        #根据单元格在Excel中的编码坐标或者数字索引坐标向单元格写入数据
        #下标1开始，参数style表示字体的颜色的名字，如red

        if coordinate is not None:
            try:
                sheet.cell(coordinate = coordinate).value = content
                if  style is not None:
                    sheet.cell(coordinate = coordinate).\
                        font = Font(color=pe.RGBDict[style])
                pe.workbook.save(pe.excelFile)
            except Exception as e:
                raise  e
        elif coordinate == None and rowno is not None and colsno is not None:
            try:
                sheet.cell(row= rowno , column= colsno).value = content
                if style:
                    sheet.cell(row = rowno,column=colsno).\
                        font = Font(color=pe.RGBDict[style])
                pe.workbook.save(pe.excelFile)
            except Exception as e:
                raise  e
        else:
            raise Exception('writecell方法与吴')

    def writecellcurrenttime(self,sheet,coordinate= None,rowno= None,colsno = None):
        #写入当前时间
        now = int(time.time())
        timearray = time.localtime(now)#显示时间戳
        currenttime = time.strftime('%Y-%m-%d %H:%M:%S',timearray)
        print(currenttime)
        #如果时间不为空
        if coordinate is not None:
            try:
                #把获取到的时间填充到对象‘coordinate’
                sheet.cell(coordinate =coordinate).value = currenttime
                pe.workbook.save(pe.excelFile)
            except Exception as e:
                raise e
        #如果传来的值不为空
        elif coordinate == None and rowno is not None and colsno is not None:
            try:
                #在对应的行和列添加时间
                sheet.cell(row=rowno, column=colsno).value = currenttime
                pe.workbook.save(pe.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception('writecellcurrenttime-方法有误')




if __name__ =='__main__':
    try:
        pe = readExcel.ParseExcel()
        ps = write()
        # ps.writecellcurrenttime()
        expath = '../textdata/demo-text.xlsx'
        ps.writecell(pe.getSheetbyindex(1))

        print('ok')
    except Exception as e:
        print(e)
        raise e
