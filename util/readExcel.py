#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2018/12/27 10:17
# @Author  : readExcel
# ##读取表格数据，返回的是一个列表，里面的数据以字典格式存放，通过索引得到字典然后取字典的值
#导入数据驱动模块xlsx
import openpyxl
#导入样式
from openpyxl.styles import Font, colors, Alignment
from untitled.DataDrivenFrameWork.config.varconfig import *
import xlrd
import  time
class ParseExcel(object):
    def __init__(self):
        #创建两个对象，初始值
        self.workbook = None
        self.excelFile = None
        self.font = Font(color = None)
        #颜色对应值
        self.RGBdic= {'red':'FFFF3030','green':'FF008B00'}
        #创建读取Excel对象方法
    def loadWoekBook(self,excelPathAndName):
        #Excel文件加入到内存，并获取wordbook对象
        try:
            self.workbook = openpyxl.load_workbook(excelPathAndName)
        except Exception as e:
            #如果发生异常，后面的程序不执行
            raise e
        #把获取到的Excel赋值新对象
        self.excelFile = excelPathAndName
        #返回Excel文件对象
        return self.workbook
    #创建方法，访问Excel的所有sheet表
    def getSheetByname(self,sheetname):
        #根据sheet名获取该sheet对象
        try:
            return self.workbook.get_sheet_by_name(sheetname)
        except Exception as e:
            raise e
    def getSheetbyindex(self,sheetindex):
        #根据sheet索引号获取该sheet对象
        try:
            #根据返回所有的sheet获取指定的索引sheet
            sheetname = self.workbook.get_sheet_names()[sheetindex]
        except Exception as e :
            print(e)
            raise e
        return self.workbook.get_sheet_by_name(sheetname)
    #获取sheet中最大行数
    def getrownumber(self,sheet):
        return sheet.max_row

    # 获取所有的最大列数
    def getcolsnumber(self,sheet):
        return sheet.max_column
    #获取sheet中最小行数
    def getstartrownumber(self,sheet):
        return sheet.min_row

    # 获取所有的最小列数
    def getstartcolsnumber(self,sheet):
        return sheet.min_column
    # def getrow(self,sheet,rowns):
    #     try:
    #         return list(sheet.rows)[rowns-1]
    #     except Exception as e:
    #         raise e
    #
    # def getColumn(self, sheet, colsno):
    #     try:
    #         return list(sheet.columns)[colsno-1]
    #     except Exception as e:
    #         raise e
    def getrow(self,sheet,nows):
        try:
            return [cell.value for cell in list(sheet.rows)[nows-1]]
        except Exception as e:
            raise e
    def getColumn(self,sheet,colsno):
        # 获取sheet某一列，返回这一列所有的数据
        # 下标1开始，sheet.rows[1]表示第一列
        try:
            return [cell.value for cell in list(sheet.columns)[colsno-1]]
        except Exception as e:
            raise  e

    def getcolrow(self,sheet,rowsno,colsno):
        #获取某行某一列
        pass

    def dict_data(self,sheet):
        # 打开excel文件
        self.data = xlrd.open_workbook(sheet)
        # 获取excel文档中第一个sheet
        self.table = self.data.sheet_by_index(1)
        # 获取第一行作为key值
        self.keys = self.table.row_values(0)

        # 获取总行数
        self.rowNum = self.table.nrows

        # 获取总列数
        self.colNum = self.table.ncols
        #获取第二行
        # 判断，如果第一个sheet行数小于1行，抛出异常，结束程序
        if self.rowNum <= 1:
            print("总行数小于1")
        else:
            # 设置空list，用来存储文件中的数据
            r = []
            # 因为第一行是标题所以总行数要-1
            for j, i in enumerate(list(range(self.rowNum - 1)), start=1):
                # 创建空对象
                s = {'rowNum': i + 2}
                # values=第二行的数据，此时j是1
                values = self.table.row_values(j)
                # 遍历列数，总共13,
                for x in list(range(self.colNum)):
                    # 让每一列keys等于values值，这样才不会乱
                    s[self.keys[x]] = values[x]
                # 因为val值赋给了keys，所以把值s拿出来放到新建的list中r
                r.append(s)
            return r

    def getcellofvalue(self,sheet,coordinate = None,rowno = None,colsno=None):
        #获取某一行某一列的数据，sheet:第几个sheet表，返回真实数据
        #根据单元格所在位置索引获取该单元格中的值，从1开始
        #sheet.cell(row=1,column=1).value,表示Excel中第一行第一列的值
        if coordinate != None:

            try:
                return sheet[coordinate].value
            except Exception as e:
                raise e
        elif coordinate is None and rowno is not None and colsno is not None:
            try:
                return sheet.cell(row = rowno,column = colsno).value

            except Exception as e:
                raise  e
        else:
            raise Exception('getcellofvalue方法有误')
    def getcellofboject(self,sheet,coordinate = None,rownoo = None,colsno = None):
        #获取某个单元格的对象，可以根据单元格所在位置的数字索引
        #也可以直接根据Excel中单元格的编码及坐标
        #如getcellboject(sheet,coordinate = 'A3') or getgetcellboject(sheet,rowno=1,colsno=2)
        if coordinate!=None:
            try:
                return sheet.cell(coordinate=coordinate)
            except Exception as e:
                raise  e
        elif rownoo is not None and colsno is not None:
            try:
                return sheet.cell(row = rownoo,column = colsno)
            except Exception as e:
                raise  e
        else:
            raise Exception('getcellofboject方法有误')

    def writecell(self, sheet, content, coordinate=None, rowno=None, colsno=None, style=None):
        # 根据单元格在Excel中的编码坐标或者数字索引坐标向单元格写入数据
        # 下标1开始，参数style表示字体的颜色的名字，如red
        # styls = Font(name='等线', size=12, color=style)
        if coordinate is not None:
            try:
                sheet.cell(coordinate=coordinate).value = content
                if style is not None:
                    sheet.cell(coordinate=coordinate).font=Font(color=self.RGBdic[style])
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif rowno is not None and colsno is not None:
            try:
                sheet.cell(row=rowno, column=colsno).value = content
                if style:
                    sheet.cell(row=rowno, column=colsno).font=Font(color=self.RGBdic[style])
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception('writecell方法与吴')

    def writecellcurrenttime(self, sheet, coordinate=None, rowno=None, colsno=None):
        # 写入当前时间
        # styls = Font(name='等线', size=12, color=style)
        now = int(time.time())
        timearray = time.localtime(now)  # 显示时间戳
        currenttime = time.strftime('%Y-%m-%d %H:%M:%S', timearray)
        print(currenttime)
        # 如果时间不为空
        if coordinate is not None:
            try:
                # 把获取到的时间填充到对象‘coordinate’
                sheet.cell(coordinate=coordinate).value = currenttime
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif rowno is not None and colsno is not None:
            try:
                # 在对应的行和列添加时间
                sheet.cell(row=rowno, column=colsno).value = currenttime

                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception('writecellcurrenttime-方法有误')
    # def writeexcel(self, sheet, content=None,  rowno=None, colsno=None, style=None):
    #         try:
    #             sheet.cell(row=rowno,column=colsno).value = content
    #
    #             self.workbook.save(self.excelFile)
    #         except Exception as e:
    #             raise e

if __name__=='__main__':
    #初始化类
    try:
        pe = ParseExcel()
        datapath = '../textdata/126邮箱创建联系人并发送邮件.xlsx'
        pe.loadWoekBook(datapath)
        # print('通过名称获取sheet对象的名字',pe.getSheetByname('Sheet1'))
        print('通过index序号获取sheet对象的名字',pe.getSheetbyindex(0).title)
        sheet = pe.getSheetbyindex(0)
        sheet1 = pe.getSheetByname('测试用例')

        print(pe.getcellofboject(sheet=sheet,coordinate="C1"))
    except Exception as e:
        raise e





